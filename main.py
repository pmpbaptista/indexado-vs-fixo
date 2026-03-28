import os
import requests
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook


# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================

MEU_CONSUMO = 700  # kWh/mês — altera consoante o teu caso

# Dados tarifários G9
POTENCIA_DIA_SMART_INDEX = 0.4236
ACESSO_REDES_SMART_INDEX = 0.0607
MARGEM_G9_ESTIMADA       = 0.0150
POTENCIA_DIA_FIXO        = 0.4498
ENERGIA_FIXO_KWH         = 0.1348

DIAS_HISTORICO = 30  # janela de cálculo da média OMIE


# ==============================================================================
# OMIE — DADOS OFICIAIS
# ==============================================================================

def obter_precos_dia_omie(data: date) -> list[float]:
    """
    Descarrega o ficheiro oficial do OMIE para uma data e devolve
    a lista de preços horários de Portugal (€/MWh).

    Fonte: https://www.omie.es  (ficheiros públicos, sem autenticação)
    Formato das linhas: ANO;MES;DIA;HORA;PRECO_PT;PRECO_ES;*
    """
    nome_ficheiro = f"marginalpdbcpt_{data.strftime('%Y%m%d')}.1"
    url = (
        "https://www.omie.es/en/file-download"
        f"?parents=marginalpdbcpt&filename={nome_ficheiro}"
    )
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer":    "https://www.omie.es/",
    }

    resposta = requests.get(url, headers=headers, timeout=15)
    resposta.raise_for_status()                          # levanta erro se 4xx/5xx
    conteudo = resposta.content.decode("latin-1")

    precos = []
    for linha in conteudo.splitlines():
        partes = linha.strip().split(";")
        # Ignora cabeçalho e linhas de totais/rodapé
        if len(partes) >= 5 and partes[0].strip().isdigit():
            try:
                precos.append(float(partes[4].replace(",", ".")))
            except ValueError:
                pass

    return precos


def obter_media_omie() -> tuple[float, str, str]:
    """
    Calcula a média OMIE (Portugal - Simples) dos últimos DIAS_HISTORICO dias,
    indo directamente à fonte oficial do OMIE.
    Devolve (media_€_MWh, data_inicio_iso, data_fim_iso).
    """
    hoje   = date.today()
    inicio = hoje - timedelta(days=DIAS_HISTORICO)

    print(f"A consultar OMIE de {inicio.strftime('%d/%m/%Y')} a {hoje.strftime('%d/%m/%Y')}...")

    todos_precos: list[float] = []
    d = inicio
    while d <= hoje:
        try:
            precos = obter_precos_dia_omie(d)
            todos_precos.extend(precos)
        except requests.HTTPError as e:
            # Dia ainda sem dados (ex.: amanhã) é normal — ignora silenciosamente
            if e.response.status_code != 404:
                print(f"  Aviso HTTP {e.response.status_code} para {d}")
        except Exception as e:
            print(f"  Aviso: não foi possível obter dados para {d} — {e}")
        d += timedelta(days=1)

    if not todos_precos:
        print("  Sem dados disponíveis. A usar valor de fallback.")
        return 65.04, inicio.isoformat(), hoje.isoformat()

    media = sum(todos_precos) / len(todos_precos)
    return media, inicio.isoformat(), hoje.isoformat()


# ==============================================================================
# HISTÓRICO EM EXCEL
# ==============================================================================

def guardar_historico_excel(data_inicio: str, data_fim: str, media: float) -> None:
    """
    Guarda o registo no ficheiro Excel ao lado do script.
    Cria o ficheiro (com cabeçalho) se ainda não existir.
    """
    nome_ficheiro  = "historico_omie.xlsx"
    diretorio      = os.path.dirname(os.path.abspath(__file__))
    caminho        = os.path.join(diretorio, nome_ficheiro)

    if not os.path.exists(caminho):
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico OMIE"
        ws.append(["Data Início", "Data Fim", "Média OMIE (€/MWh)"])
    else:
        wb = load_workbook(caminho)
        ws = wb.active

    ws.append([data_inicio, data_fim, media])
    wb.save(caminho)
    print(f"💾 Registo guardado: {caminho}")


# ==============================================================================
# ANÁLISE DO PONTO DE EQUILÍBRIO G9
# ==============================================================================

def analisar_ponto_equilibrio_g9(consumo_mensal: float) -> None:

    # --- Média OMIE (fonte oficial, sem browser) ---
    media_omie, data_inicio, data_fim = obter_media_omie()

    # --- Guardar histórico ---
    guardar_historico_excel(data_inicio, data_fim, media_omie)

    # --- Cálculos ---
    custo_fixo_total   = (consumo_mensal * ENERGIA_FIXO_KWH) + (DIAS_HISTORICO * POTENCIA_DIA_FIXO)
    termo_potencia_smart = DIAS_HISTORICO * POTENCIA_DIA_SMART_INDEX

    # Preço-limite OMIE (break-even) em €/MWh
    omie_limite = (
        (custo_fixo_total - termo_potencia_smart) / consumo_mensal
        - MARGEM_G9_ESTIMADA
        - ACESSO_REDES_SMART_INDEX
    ) * 1000

    # Custo estimado no tarifário indexado
    preco_kwh_smart   = (media_omie / 1000) + MARGEM_G9_ESTIMADA + ACESSO_REDES_SMART_INDEX
    custo_smart_total = (consumo_mensal * preco_kwh_smart) + termo_potencia_smart

    # --- Relatório ---
    print(f"\n{'#'*60}")
    print(f"  ANÁLISE DE MERCADO OMIE (Últimos {DIAS_HISTORICO} dias)")
    print(f"{'#'*60}")
    print(f"Média OMIE extraída:          {media_omie:>8.2f} €/MWh")
    print(f"Seu Ponto de Equilíbrio:      {omie_limite:>8.2f} €/MWh")
    print(f"{'-'*60}")
    print(f"Custo ESTIMADO INDEXADO:      {custo_smart_total:>8.2f} €")
    print(f"Custo FIXO (Vantagem+):       {custo_fixo_total:>8.2f} €")
    print(f"{'-'*60}")

    if media_omie > omie_limite:
        print(f"🚨 ALERTA: A média OMIE está ALTA ({media_omie:.2f} > {omie_limite:.2f}).")
        print("   O tarifário FIXO é atualmente mais vantajoso.")
    else:
        print(f"✅ POUPANÇA: A média OMIE está BAIXA ({media_omie:.2f} < {omie_limite:.2f}).")
        print("   O tarifário INDEXADO é a melhor escolha agora.")

    print(f"{'#'*60}\n")


# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    analisar_ponto_equilibrio_g9(MEU_CONSUMO)